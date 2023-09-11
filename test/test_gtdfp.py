import os

import pandas
from openpyxl import load_workbook

from sources.fine_reader_module.pywinauto_fr import recognition_using_fine_reader
from sources.settings import config_class
from sources.recognition_module.pandas_work_with_file import delete_temp_files, get_width_column, resize_column, \
    write_finereader_data_to_excel, convert_data_list_in_excel, recognition_main

config_class.poppler_config.poppler_path = os.path.abspath(r'..\utils\poppler-23.01.0\Library\bin')
config_class.path_project.temp_dir = os.path.abspath(r'temp_test')


# pytest -rA -v --tb=line test_gtdfp.py --cov-report term-missing --cov=sources

class TestWorkWithFile:

    def test_delete_temp_files(self):
        test_path = r'temp_test\test.txt'
        with open(test_path, 'w') as file:
            file.write('Тест')
        delete_temp_files(test_path)
        assert not os.path.exists(test_path)

    def test_get_width_column(self):
        df_dict = {'Column_1': ['test_1', 'test_2', 'test_3'],
                   'Column_2': ['test_4', 'test_5', 'test_6']}
        df_test = pandas.DataFrame(df_dict)
        result = get_width_column(df_test)
        assert all([isinstance(result, list), (len(result) == 2), (max(result) == 6)])

    def test_resize_column(self):
        df_dict = {'Column_1': ['more characters to test', 'test_2', 'test_3'],
                   'Column_2': ['test_4', 'test_5', 'test_test_test_test']}
        df_test = pandas.DataFrame(df_dict)
        test_path = r'temp_test\test_excel_resize.xlsx'
        df_test.to_excel(test_path)
        sheet_name = 'Лист1'
        width_list_1 = get_width_column(df_test)
        resize_column(test_path, width_list_1, sheet_name)
        df_test_2 = pandas.read_excel(test_path)
        width_list_2 = get_width_column(df_test_2)
        os.remove(test_path)
        assert width_list_1 != width_list_2

    def test_write_finereader_data_to_excel(self):
        test_fr_path = r'temp_test\test_excel_fr.xlsx'
        test_path = r'temp_test\test_excel.xlsx'
        write_finereader_data_to_excel(test_path, test_fr_path)
        workbook = load_workbook(test_path)
        assert 'Данные FineReader' in [sheet for sheet in workbook.sheetnames]

    def test_convert_data_list_in_excel(self):
        data_list = [['test_1', 'test_2', 'test_3'], ['test_4', 'test_5', 'test_6']]
        test_path = r'temp_test\test_excel.xlsx'
        convert_data_list_in_excel(data_list, test_path)
        workbook = load_workbook(test_path)
        assert 'Данные Tesseract' in [sheet for sheet in workbook.sheetnames]

    def test_recognition_main(self):
        test_path = os.path.abspath(r'temp_test\PDF-test.pdf')
        recognition_main(test_path, 0)
        result_path = r'temp_test\PDF-test.xlsx'
        workbook = load_workbook(result_path)
        sheet_tuple = ('Данные FineReader', 'Данные Tesseract')
        assert all([os.path.exists(result_path), all([i in [sh for sh in workbook.sheetnames] for i in sheet_tuple])])
        os.remove(result_path)


class TestFineReader:
    def test_recognition_using_fine_reader(self):
        test_path = os.path.abspath(r'temp_test\PDF-test.pdf')
        save_path = os.path.abspath(r'temp_test\PDF-test.xlsx')
        recognition_using_fine_reader(test_path, save_path)
        assert os.path.exists(save_path)
        os.remove(save_path)
