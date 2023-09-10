import logging
import os

import pandas
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from sources.fine_reader_module.pywinauto_fr import recognition_using_fine_reader
from sources.tesseract_module.ocr_tesseract import RecognitionModule

logger = logging.getLogger(__name__)


def recognition_main(pdf_file_path: str, degree_of_rotation: int) -> str:
    """
    Функция основной логики запуска распознавания файла.
    Получаем данные из tesseract_module, далее получаем данные из fine_reader_module - объединяем данные в один
    результирующий excel-файл, который потом пойдет на отправку.

    :param pdf_file_path: Путь до полученого PDF-файла
    :param degree_of_rotation: Наклон изображения
    :return: Путь до результирующего excel-файла
    """
    # Шаг №1 - произведем распознавание с помощью Тессеракта
    ocr_obj = RecognitionModule(pdf_file_path, degree_of_rotation)
    data_matrix = ocr_obj.data_from_pdf

    # Шаг №2 - сохраним полученные данные в excel
    path_to_save = pdf_file_path.split('.')[0] + '.xlsx'
    convert_data_list_in_excel(data_matrix, path_to_save)

    # Шаг №3 - произведем распознавание с помощью FineReader
    path_to_save_fr = os.path.abspath(pdf_file_path.split('.')[0] + '_fr.xlsx')
    logger.info(f'Абсолютный путь: {path_to_save_fr}')
    recognition_using_fine_reader(os.path.abspath(pdf_file_path), path_to_save_fr)

    # Шаг №4 - сохраним данные в общий excel
    write_finereader_data_to_excel(path_to_save, path_to_save_fr)

    # Шаг №5 - удалим файл excel файл FineReader
    os.remove(path_to_save_fr)
    return path_to_save


def convert_data_list_in_excel(data_matrix: list, path_to_save: str) -> str:
    """
    Функция конвертирования полученных данных с Tesseract, в Excel - файл, дополнительно проводим выравнием столбцов
    по ширине содержимого в ячейках

    :param data_matrix: Данные полученные при распознавании файла, с oct_tesseract.py
    :param path_to_save: Путь до результирующего excel-файла
    :return: Конечный путь, использовать при необходимости проверки на None
    """
    sheet_name = 'Данные Tesseract'

    try:
        # Преобразуем матрицу в DataFrame и запишем её на лист
        data_dict = dict(zip(range(1, len(data_matrix) + 1), data_matrix))
        df = pandas.DataFrame.from_dict(data_dict, orient='index')
        df.insert(1, "№", df.index + 1)
        logger.info('Сформировал DF из распознанных данных')
        with pandas.ExcelWriter(path_to_save) as writer:
            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, engine='openpyxl')
        # Выравниваем столбцы по содержимому ячеек
        width_list_data = get_width_column(df)
        if width_list_data:
            resize_column(path_to_save, width_list_data, sheet_name)
        return path_to_save
    except Exception as e:
        logger.error(f'Ошибка в функции convert_data_list_in_excel - {e}')
        raise e


def write_finereader_data_to_excel(path_to_save: str, path_to_save_fr: str) -> str:
    """
    Функция записи полученных данных из FineReader, в результируюх excel-файл на другой лист.
    Дополнительно проводим выравнием столбцов по ширине содержимого в ячейках

    :param path_to_save: Путь до результирующего excel-файла
    :param path_to_save_fr: Путь, до полученного excel-файла из FineReader
    :return: Конечный путь, использовать при необходимости проверки на None
    """
    sheet_name = 'Данные FineReader'
    try:
        df_fr = pandas.read_excel(path_to_save_fr)
        with pandas.ExcelWriter(path_to_save, mode="a", if_sheet_exists="replace") as writer:
            df_fr.to_excel(writer, sheet_name=sheet_name, header=False, index=False, engine='openpyxl')
        # Выравниваем столбцы по содержимому ячеек
        width_list_data = get_width_column(df_fr)
        if width_list_data:
            resize_column(path_to_save_fr, width_list_data, sheet_name)
        return path_to_save
    except Exception as e:
        logger.error(f'Ошибка в функции write_finereader_data_to_excel - {e}')
        raise e


def resize_column(path: str, width_list: list, sheetname: str) -> None:
    """
    Функция, которая устанавливает новую ширину столбцов таблицы, исходя из переданного списка.

    :param path: Путь до excel-файла
    :param width_list: Список ширины столбцов таблицы
    :param sheetname: Имя листа excel-файла, на котором будем проводить форматирование
    :return: None
    """
    try:
        wb = load_workbook(path)
        ws = wb[sheetname]
        ws.column_dimensions['A'].width = 5
        for i, j in zip(range(1, ws.max_column + 1), width_list):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = j + 2
        wb.save(path)
    except Exception as e:
        logger.error(f'Ошибка при изменении ширины столбцов в excel - {e}')


def get_width_column(df: pandas.DataFrame) -> list:
    """
    Функция получения ширины столбца таблицы, по максимальной длиней, содержащихся в ячейки данных.

    :param df: pandas.Dataframe
    :return: Список ширины столбцов таблицы
    """
    try:
        df = df.astype(str)
        width_list = [max([len(i) if i is not None else 0 for i in tuple(df[column])]) for column in
                      tuple(df.columns)]
        return width_list
    except Exception as e:
        logger.error(f'Ошибка при получении ширины столбцов - {e}')


def delete_temp_files(*agrs: str) -> None:
    """
    Функция для удаления файлов

    :param agrs: Пути до файлов, которых необходимо удалить
    :return: None
    """
    try:
        del_list = [i for i in agrs if isinstance(i, str) and os.path.exists(i)]
        for file in del_list:
            os.remove(file)
        logger.info(f'Удалил временные файлы: {del_list}')
    except Exception as e:
        logger.error(f'Ошибка при удалении временных файлов: {e}')
